# Ch02. 쇼핑몰 주문 요청서 분류 자동화 프로젝트 - 05. 분류한 파일에 주문건수 넣기
import pandas as pd
from openpyxl.reader.excel import load_workbook
from datetime import datetime
# pd.set_option('display.max_columns', None)

class ClassificationExcel:

    path = ''

    def __init__(self, order_xlsx_filename, partner_info_xlsx_filename, path='result'):
        # 주문목록 (데이터를 가공함)
        df = pd.read_excel(order_xlsx_filename)
        df = df.rename(columns=df.iloc[1]) #1번을 열제목으로 쓰겠다
        df = df.drop([df.index[0], df.index[1]])
        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path

        # 파트너목록
        df_partners = pd.read_excel(partner_info_xlsx_filename)

        self.brands = df_partners['브랜드'].to_list()
        self.partners = df_partners['업체명'].to_list()


    def classify(self):
        
        for i, row in self.order_list.iterrows(): # df 데이터를 >> 이터로우즈=2개의 행들을 쭉 반복하겠다는 뜻 / head(5)=너무 많아서 복잡하니까 위에서 5개만 뽑겠다
            brand_name = ''
            partner_name = ''
            idx_partner = 0
            for j in range(len(self.brands)):
                if self.brands[j] in row['상품명']:

                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break
            # print(f'{row["상품명"]} 은 {brand_name} 브랜드 입니다. {j}번째')
            # print(f'업체명:{partner_name}')

            df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
            # print(df_filtered)
            df_filtered.to_excel(f'{self.path}/[메가몰] {partner_name}.xlsx')

    def set_count(self):
        file_name = '20250117/[메가몰] 다온마켓.xlsx'
        wb = load_workbook(file_name)
        ws = wb.active
        print('value:', ws['B1'].value)
        print('value:', ws['B2'].value)

        # 개수 세기
        row_cnt = ws.max_row - 1
        print('cnt:', row_cnt)

        # 열 삽입
        ws.insert_rows(1)
        ws.insert_rows(1)

        now_day = datetime.now().strftime('%Y-%m-%d')
        # A1
        ws['A1'] = f'발송요청내역 [총 {row_cnt}건] {now_day}'


        wb.save(file_name)


if __name__ == '__main__':
    ce = ClassificationExcel('주문목록20221112_NEW.xlsx', '파트너목록_NEW.xlsx', '20250117')
    # ce.classify()
    ce.set_count()